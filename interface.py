#!/usr/local/bin python3

from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from ttkthemes import ThemedTk

from create_sheets import *

# generate sheet datos
def generate_data(default_route, default_name, default_sheet): 
  global ndf
  
  route=th.get()
  sheet=ts.get()
  
  if len(route) == 0:
    route=default_route+default_name
    th.insert(0,route)
  if len(sheet) == 0:
    sheet=default_sheet
    ts.insert(0,sheet)
  
  sroute=route.split("/")
  namef=sroute[-1].strip()
  route="/".join(sroute[:-1])+"/"
  if route == "/":
    route="."+route
  print(route,namef)
  
  try:
    df,route,sheet,nsheet=read_sheet(loc=route,sheet=sheet,namef=namef)
    df = create_data(df)
    
    wb = load_workbook(route,data_only=False)
    writer = pd.ExcelWriter(route,engine='openpyxl')
    writer.book = wb
    
    try:
      writer.book.remove(writer.book[nsheet])
    except:
      messagebox.showinfo("Datos","Hoja %s no existe. Creando..."%(nsheet))
    finally:
      df.to_excel(writer,sheet_name=nsheet,index=False) 
      writer.save()
    writer.close()
    
    messagebox.showinfo("Datos","%s\nArchivo creado\nRuta: %s\nHoja: %s"%(namef,route,nsheet))
    labelb.configure(text="Datos cargados en hoja Datos (libro %s)"%(namef))
    labelg.configure(text="No hay datos cargados para generar las graficas")
    
  except:
    messagebox.showerror("Datos","No se localizo el archivo.")
  ndf = df
  return ndf
  
#generate sheet base en miles
def generate_base_mil(nsheet):
  global graph_df,rates_df
  
  route=th.get()
  
  df,rdf=base_mil(ndf)
  graph_df,rates_df=df,rdf
  
  namef=route.split("/")[-1].strip()
  route2="/".join(route.split("/")[:-1])+"/"  
  if route2 == "/":
    route2="."+route2
  route=route2+namef
  
  try:
    wb = load_workbook(route,data_only=False)
    writer = pd.ExcelWriter(route,engine='openpyxl')
    writer.book = wb  
    
    try:
      writer.book.remove(writer.book[nsheet])
    except:
      messagebox.showinfo("Base Mil","Hoja %s no existe. Creando..."%(nsheet))
    finally:
      df.to_excel(writer,sheet_name=nsheet,
                  index=False,startrow=0,startcol=0)
      rdf.to_excel(writer,sheet_name=nsheet,
                   index=True,startrow=21,startcol=1)
      writer.save()
    writer.close()
    
    messagebox.showinfo("Base Mil","%s\nArchivo creado\nRuta: %s\nHoja: %s"%(namef,route2,nsheet))
    labelg.configure(text="Datos cargados en hoja Base Mil (libro %s)"%(namef))
  except:
    messagebox.showerror("Base Mil","No se localizo el archivo")
  return graph_df,rates_df

# function to plot the data in base mil
def generate_graphs():
  fig,axs=plt.subplots(3,1,gridspec_kw={'hspace':0.4,'wspace':0.7},
                       figsize=(8,8))
  
  #=======================================#
  # graph for total
  
  x=graph_df.columns[2:].tolist()
  y=graph_df.iloc[-1,2:]
  z=np.insert(np.array(rates_df.iloc[0,1:]),0,0)*100
  p1=axs[0].plot(x,y,
              'k-',label="Cantidad Total")
  axs[0].set_xlabel("Dias")
  axs[0].set_ylabel("Puntos redimidos en miles")
  axs[0].set_xlim(min(x),max(x))
  axs[0].set_xticks(range(len(x)+1))
  axs[0].set_ylim(0,max(y))
  axs[0].set_title("Redencion total")
  
  axs0=axs[0].twinx()
  p2=axs0.plot(x,z,'k--',label='Rate')
  axs0.set_ylabel("Rate (%)")
  axs0.set_ylim(min(z),max(z))
  
  # added these three lines
  lns = p1+p2
  labs = [l.get_label() for l in lns]
  axs[0].legend(lns, labs, loc="best")
  #=======================================#
  #graph for online
  
  x=graph_df.columns[2:].tolist()
  y=graph_df.iloc[12,2:]
  z=np.insert(np.array(rates_df.iloc[1,1:]),0,0)*100
  axs[1].plot(x,y,
              'k-',label="Cantidad Total")
  axs[1].set_xlabel("Dias")
  axs[1].set_ylabel("Puntos redimidos en miles")
  axs[1].set_xlim(min(x),max(x))
  axs[1].set_xticks(range(len(x)+1))
  axs[1].set_ylim(0,max(y))
  axs[1].set_title("Tiendas online (Linio)")
  
  axs0=axs[1].twinx()
  axs0.plot(x,z,'k--',label='Rate')
  axs0.set_ylabel("Rate (%)")
  axs0.set_ylim(min(z),max(z))
  #=======================================#
  # graph for aereo 
  
  x=graph_df.columns[2:].tolist()
  y=graph_df.iloc[1,2:]
  z=np.insert(np.array(rates_df.iloc[2,1:]),0,0)*100
  axs[2].plot(x,y,
              'k-',label="Cantidad Total")
  axs[2].set_xlabel("Dias")
  axs[2].set_ylabel("Puntos redimidos en miles")
  axs[2].set_xlim(min(x),max(x))
  axs[2].set_xticks(range(len(x)+1))
  axs[2].set_ylim(0,max(y))
  axs[2].set_title("Tiendas aereo")
  
  axs0=axs[2].twinx()
  axs0.plot(x,z,'k--',label='Rate')
  axs0.set_ylabel("Rate (%)")
  axs0.set_ylim(min(z),max(z))
  #=======================================#
  
  plt.show()
  plt.clf()

# load and save data in curr
def save_data(default_route):
  route=th.get()
  
  if len(route) == 0:
    route=default_route
  
  try:  
    wbs = xw.Book(route)
    wbs.save(default_route)
    wbs.close()
    wbs.app.quit()
    messagebox.showinfo("Datos", "Datos guardados")
  except:
    messagebox.showerror("Datos", "No se encontro archivo")
    wbs.app.quit()

def open_excel():
  route=of.get()
  try:
    wb=xw.Book(route)
  except:
    messagebox.showerror("Abrir archivo", "Archivo no encontrado")
  
###########################################

# GUI code

def_route="/Users/alexlopez/Documents/Documentos-varios/automation_tests/"
def_name="e1.xlsx"
def_sheet="Base"    

win=ThemedTk(theme="adapta") #creating the main window and storing the window object in 'win' 
win.geometry("540x400")

#======================================================#
padx=10
pady=10
lf=ttk.LabelFrame(win,text="Datos",width=500, height=100)
lf.grid(column=0,row=0,padx=padx,pady=pady)
lf.grid_propagate(0)

ldirh=ttk.Label(lf, text='Ruta').grid(row=0) 
th=ttk.Entry(lf)
th.grid(row=0,column=1)
nsheeth=ttk.Label(lf, text='Sheet').grid(row=0,column=2) 
ts=ttk.Entry(lf)
ts.grid(row=0,column=3)

btnh=ttk.Button(lf,text="Ejecutar datos", width=12,
            command=lambda: generate_data(def_route,def_name,def_sheet))
btnh.grid(row=1,column=3)

#======================================================#
lb=ttk.LabelFrame(win,text="Base Mil",width=500, height=100)
lb.grid(column=0,row=2,padx=padx,pady=pady)
lb.grid_propagate(0)

labelb=ttk.Label(lb, text='No hay datos cargados para Base Mil')
labelb.grid(row=2)

btnb=ttk.Button(lb,text="Ejecutar base mil", width=12,
            command=lambda: generate_base_mil("Base en miles"))
btnb.place(x=292,y=25)

#======================================================#
lg=ttk.LabelFrame(win,text="Graficas",width=500, height=100)
lg.grid(column=0,row=4,padx=padx,pady=pady)
lg.grid_propagate(0)

labelg=ttk.Label(lg, text='No hay datos cargados para generar las graficas')
labelg.grid(row=2)

btng=ttk.Button(lg,text="Ejecutar graficas",
            command=generate_graphs)
btng.place(x=292,y=25)

#======================================================#
btnsave=ttk.Button(win,text="Guardar datos", width=12,
            command=lambda: save_data(def_route+def_name))
btnsave.place(x=390,y=360)

#======================================================#
labelo=ttk.Label(win, text='Ruta')
labelo.place(x=10,y=365)

of=ttk.Entry(win)
of.place(x=45,y=363)

btnopen=ttk.Button(win,text="Abrir archivo",width=12,
            command=lambda: open_excel())
btnopen.place(x=235,y=360)

win.mainloop()